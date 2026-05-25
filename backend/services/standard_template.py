"""Standard EBNA/EBDT torque template support.

The Excel workbook is the source of truth for printed fields. OCR should still
read handwritten values from uploaded PDFs, but operation/process/quantity/spec
values are corrected from these template rows before saving.
"""

from __future__ import annotations

import logging
import json
import os
import re
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import StandardTemplateRow

logger = logging.getLogger(__name__)

BACKEND_DIR = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE_PATH = BACKEND_DIR / "data" / "standard_torque_template.xlsx"
DEFAULT_RULES_PATH = BACKEND_DIR / "data" / "template_rules.json"

TORQUE_SHEETS = {
    "Torque Audit Sheet - EBNA": {
        "model": "EBNA",
        "start": 5,
        "op": 1,
        "process": 2,
        "tightening_equipment": 6,
        "quantity": 8,
        "tightening_torque": 9,
        "checking_equipment": 11,
    },
    "Torque Audit Sheet - EBDT": {
        "model": "EBDT",
        "start": 5,
        "op": 1,
        "process": 3,
        "tightening_equipment": 7,
        "tightening_part": 9,
        "quantity": 10,
        "tightening_torque": 11,
        "engineering_spec": 13,
        "checking_equipment": 14,
    },
}


def _excluded_ops() -> set[str]:
    rules = _template_rules()
    configured = rules.get("excluded_ops") or []
    env_raw = os.getenv("EXCLUDED_TEMPLATE_OPS")
    env_values = re.split(r"[,;\s]+", env_raw) if env_raw else []
    values = [*configured, *env_values]
    return {re.sub(r"\D", "", str(item)) for item in values if re.sub(r"\D", "", str(item))}


def default_template_model() -> str:
    return os.getenv("DEFAULT_TEMPLATE_MODEL", str(_template_rules().get("default_model") or "EBDT")).upper()


@lru_cache(maxsize=1)
def _template_rules() -> dict:
    configured = os.getenv("STANDARD_TEMPLATE_RULES_JSON")
    path = Path(configured).resolve() if configured else DEFAULT_RULES_PATH
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("Could not read standard template rules from %s: %s", path, exc)
        return {}


def _quantity_override(model: str, operation_number: str, sequence: int) -> int | None:
    rules = _template_rules().get("quantity_overrides") or {}
    model_rules = rules.get(model) or rules.get(model.upper()) or {}
    values = model_rules.get(operation_number) or model_rules.get(str(operation_number))
    if not isinstance(values, list) or sequence < 1 or sequence > len(values):
        return None
    try:
        return max(0, int(float(values[sequence - 1])))
    except (TypeError, ValueError):
        return None


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r", "\n").replace("ą", "+/-").replace("±", "+/-")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def _clean_op(value) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    parts = re.findall(r"\d+", text)
    if len(parts) == 1 and len(parts[0]) > 4 and len(parts[0]) % 4 == 0:
        parts = [parts[0][index: index + 4] for index in range(0, len(parts[0]), 4)]
    return "&".join(parts) if parts else None


def _template_op_text(value) -> str | None:
    text = _clean_text(value)
    return text if _clean_op(text) else None


def _op_aliases(operation_number: str | None) -> set[str]:
    op = _clean_op(operation_number)
    if not op:
        return set()
    parts = op.split("&")
    aliases = {op, "".join(parts)}
    if len(parts) > 1:
        aliases.update(parts)
    return {alias for alias in aliases if alias}


def _clean_quantity(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return None


def _template_path() -> Path:
    configured = os.getenv("STANDARD_TEMPLATE_XLSX")
    return Path(configured).resolve() if configured else DEFAULT_TEMPLATE_PATH


def _extract_rows_from_workbook(path: Path) -> list[dict]:
    if not path.exists():
        logger.warning("Standard template workbook not found: %s", path)
        return []

    workbook = load_workbook(path, data_only=True, read_only=True)
    extracted: list[dict] = []

    for sheet_name, spec in TORQUE_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            logger.warning("Standard template sheet missing: %s", sheet_name)
            continue

        sheet = workbook[sheet_name]
        current_op: str | None = None
        current_op_key: str | None = None
        current_process: str | None = None
        current_equipment: str | None = None
        sequence_by_op: dict[str, int] = {}

        for row_num in range(spec["start"], sheet.max_row + 1):
            op_cell = sheet.cell(row_num, spec["op"]).value
            op_key = _clean_op(op_cell)
            if op_key:
                current_op = _template_op_text(op_cell) or op_key
                current_op_key = op_key
                current_process = _clean_text(sheet.cell(row_num, spec["process"]).value) or current_process
                current_equipment = (
                    _clean_text(sheet.cell(row_num, spec["tightening_equipment"]).value)
                    or current_equipment
                )

            quantity = _clean_quantity(sheet.cell(row_num, spec["quantity"]).value)
            if not current_op or not current_op_key or current_op_key in _excluded_ops() or quantity is None:
                continue

            process_name = _clean_text(sheet.cell(row_num, spec["process"]).value) or current_process
            tightening_equipment = (
                _clean_text(sheet.cell(row_num, spec["tightening_equipment"]).value)
                or current_equipment
            )
            row = {
                "model": spec["model"],
                "sheet_name": sheet_name,
                "operation_number": current_op,
                "sequence": sequence_by_op.get(current_op_key, 0) + 1,
                "process_name": process_name,
                "tightening_equipment": tightening_equipment,
                "tightening_part": _clean_text(sheet.cell(row_num, spec.get("tightening_part", 0)).value)
                if spec.get("tightening_part")
                else None,
                "quantity": quantity,
                "tightening_torque": _clean_text(sheet.cell(row_num, spec["tightening_torque"]).value),
                "engineering_spec": _clean_text(sheet.cell(row_num, spec.get("engineering_spec", 0)).value)
                if spec.get("engineering_spec")
                else None,
                "checking_equipment": _clean_text(sheet.cell(row_num, spec["checking_equipment"]).value),
                "source_row": row_num,
            }
            override_quantity = _quantity_override(row["model"], current_op_key, row["sequence"])
            if override_quantity is not None:
                row["quantity"] = override_quantity
            if any(row.get(k) for k in ("process_name", "tightening_torque", "engineering_spec")):
                sequence_by_op[current_op_key] = row["sequence"]
                extracted.append(row)

    return extracted


def seed_standard_templates(db: Session, force: bool = False) -> int:
    """Seed template rows from the standard workbook if the DB table is empty."""
    if force:
        db.query(StandardTemplateRow).delete(synchronize_session=False)
        db.commit()

    excluded_ops = _excluded_ops()
    if excluded_ops:
        deleted = (
            db.query(StandardTemplateRow)
            .filter(StandardTemplateRow.operation_number.in_(excluded_ops))
            .delete(synchronize_session=False)
        )
        if deleted:
            logger.info("Removed %d excluded standard template rows: %s", deleted, ", ".join(sorted(excluded_ops)))
            db.commit()

    existing = db.query(StandardTemplateRow).count()
    if existing and not force:
        return existing

    rows = _extract_rows_from_workbook(_template_path())
    for row in rows:
        db.add(StandardTemplateRow(**row))
    db.commit()
    logger.info("Seeded %d standard template rows", len(rows))
    _cached_workbook_rows.cache_clear()
    return len(rows)


def _row_to_dict(row: StandardTemplateRow) -> dict:
    return {
        "model": row.model,
        "sheet_name": row.sheet_name,
        "operation_number": row.operation_number,
        "sequence": row.sequence,
        "process_name": row.process_name,
        "tightening_equipment": row.tightening_equipment,
        "tightening_part": row.tightening_part,
        "quantity": row.quantity,
        "tightening_torque": row.tightening_torque,
        "engineering_spec": row.engineering_spec,
        "checking_equipment": row.checking_equipment,
        "source_row": row.source_row,
    }


def _load_db_template_rows(db: Session) -> list[dict]:
    rows = (
        db.query(StandardTemplateRow)
        .order_by(StandardTemplateRow.model, StandardTemplateRow.id)
        .all()
    )
    if rows:
        return _dedupe_template_rows(_row_to_dict(row) for row in rows)

    seed_standard_templates(db)
    rows = (
        db.query(StandardTemplateRow)
        .order_by(StandardTemplateRow.model, StandardTemplateRow.id)
        .all()
    )
    return _dedupe_template_rows(_row_to_dict(row) for row in rows)


def _dedupe_template_rows(rows: Iterable[dict]) -> list[dict]:
    deduped: list[dict] = []
    seen: set[tuple] = set()
    for row in rows:
        key = (
            row.get("model"),
            _clean_op(row.get("operation_number")),
            row.get("sequence"),
            row.get("process_name"),
            row.get("quantity"),
            row.get("source_row"),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


@lru_cache(maxsize=1)
def _cached_workbook_rows() -> tuple[dict, ...]:
    return tuple(_extract_rows_from_workbook(_template_path()))


def template_prompt_context(max_rows: int = 140) -> str:
    """Compact template context for the OCR prompt."""
    rows = list(_cached_workbook_rows())[:max_rows]
    if not rows:
        return ""
    lines = [
        "STANDARD PRINTED TORQUE TEMPLATE (use for op/process/quantity/spec; read handwritten Actual values from image):"
    ]
    for row in rows:
        part = f";part={row['tightening_part']}" if row.get("tightening_part") else ""
        lines.append(
            f"{row['model']}|op={row['operation_number']}|seq={row['sequence']}|qty={row['quantity']}"
            f"|process={row.get('process_name') or ''}{part}|spec={row.get('engineering_spec') or row.get('tightening_torque') or ''}"
        )
    return "\n".join(lines)


def _choose_model(ocr_rows: Iterable[dict], template_rows: list[dict], preferred_model: str | None = None) -> str | None:
    if preferred_model:
        preferred = preferred_model.upper()
        if any(row["model"] == preferred for row in template_rows):
            return preferred

    ocr_ops = [_clean_op(row.get("operation_number")) for row in ocr_rows]
    ocr_ops = [op for op in ocr_ops if op]
    available_models = {row["model"] for row in template_rows}
    fallback = default_template_model()
    if not ocr_ops:
        return fallback if fallback in available_models else next(iter(available_models), None)

    scores: dict[str, int] = {}
    for model in available_models:
        model_ops = {row["operation_number"] for row in template_rows if row["model"] == model}
        scores[model] = sum(1 for op in ocr_ops if op in model_ops)
    if scores and max(scores.values()) > 0:
        return max(scores, key=scores.get)
    return fallback if fallback in available_models else next(iter(available_models), None)


def _parse_limits(text: str | None) -> tuple[float | None, float | None, float | None]:
    if not text:
        return None, None, None
    normalized = (
        text.replace(",", ".")
        .replace("=<", "<=")
        .replace("=>", ">=")
        .replace("≤", "<=")
        .replace("≥", ">=")
    )
    range_matches = re.findall(
        r"(-?\d+(?:\.\d+)?)\s*(?:nm|kg|kn|c|°c|deg|degree|degrees)?\s*(?:~|–|—|-|to)\s*"
        r"(-?\d+(?:\.\d+)?)\s*(?:nm|kg|kn|c|°c|deg|degree|degrees)?",
        normalized,
        flags=re.I,
    )
    lower = upper = None
    if range_matches:
        a, b = range_matches[-1]
        first, second = float(a), float(b)
        lower, upper = min(first, second), max(first, second)

    torque_matches = re.findall(
        r"(?:final\s+torque|final|torque|final\s+tight)\s*:?\s*(-?\d+(?:\.\d+)?)",
        normalized,
        flags=re.I,
    )
    torque_match = torque_matches[-1] if torque_matches else None
    nominal = float(torque_match) if torque_match else None
    if nominal is not None and (lower is None or upper is None):
        tolerance_match = re.search(r"(?:\+/-|±)\s*(\d+(?:\.\d+)?)\s*%", normalized, flags=re.I)
        if tolerance_match:
            tolerance = float(tolerance_match.group(1)) / 100
            lower = nominal * (1 - tolerance)
            upper = nominal * (1 + tolerance)
    if lower is None:
        lower_match = re.search(
            r"(?:above|minimum|min|>=|=>)\s*:?\s*-?\s*(\d+(?:\.\d+)?)|"
            r"(\d+(?:\.\d+)?)\s*(?:nm|kg|kn|c|°c)?\s*(?:minimum|min)\b",
            normalized,
            flags=re.I,
        )
        if lower_match:
            lower = float(next(group for group in lower_match.groups() if group is not None))
    if upper is None:
        upper_match = re.search(
            r"(?:below|maximum|max|<=|=<)\s*:?\s*-?\s*(\d+(?:\.\d+)?)|"
            r"(\d+(?:\.\d+)?)\s*(?:nm|kg|kn|c|°c)?\s*(?:maximum|max)\b",
            normalized,
            flags=re.I,
        )
        if upper_match:
            upper = float(next(group for group in upper_match.groups() if group is not None))
    if nominal is None:
        first_number = re.search(r"-?\d+(?:\.\d+)?", normalized)
        nominal = float(first_number.group(0)) if first_number else None
    if nominal is not None and nominal < 0 and (lower == abs(nominal) or upper == abs(nominal)):
        nominal = abs(nominal)
    if lower is not None and upper is not None and upper < lower:
        upper = None
    return nominal, lower, upper


def _parse_template_limits(template: dict) -> tuple[float | None, float | None, float | None]:
    candidates = []
    for text in (template.get("tightening_torque"), template.get("engineering_spec")):
        nominal, lower, upper = _parse_limits(text)
        score = int(lower is not None) + int(upper is not None)
        if score:
            candidates.append((score, nominal, lower, upper))
    if candidates:
        _score, nominal, lower, upper = max(candidates, key=lambda item: item[0])
        return nominal, lower, upper
    return _parse_limits(template.get("engineering_spec") or template.get("tightening_torque"))


def _template_key(template: dict | None) -> tuple | None:
    if not template:
        return None
    return (
        template.get("model"),
        template.get("sheet_name"),
        template.get("source_row"),
        _clean_op(template.get("operation_number")),
        template.get("sequence"),
        template.get("process_name"),
    )


def _template_enriched_row(template: dict, base: dict | None, model: str, measurements: list[float] | None = None) -> dict:
    source = base or {}
    nominal, lower, upper = _parse_template_limits(template)
    confidence_scores = {**(source.get("confidence_scores") or {})}
    confidence_scores.update({
        "operation_number": 1.0,
        "process_name": 1.0,
        "quantity": 1.0,
    })
    return {
        **source,
        "operation_number": template["operation_number"],
        "process_name": template.get("process_name"),
        "process_description": template.get("process_name"),
        "quantity": template.get("quantity"),
        "nominal": nominal,
        "upper_limit": upper,
        "lower_limit": lower,
        "measurements": measurements or [],
        "confidence_scores": confidence_scores,
        "template": template,
        "template_model": model,
        "printed_values_source": "standard_template",
    }


def _fill_template_gaps(corrected: list[dict], template_rows: list[dict], model: str, max_gap_rows: int = 12) -> list[dict]:
    if len(corrected) < 2:
        return corrected

    model_templates = [
        row for row in template_rows
        if row.get("model") == model and row.get("source_row") is not None
    ]
    by_sheet = {}
    for template in model_templates:
        by_sheet.setdefault(template.get("sheet_name"), []).append(template)
    for sheet_rows in by_sheet.values():
        sheet_rows.sort(key=lambda row: (row.get("source_row") or 0, row.get("sequence") or 0))

    existing_keys = {_template_key(row.get("template")) for row in corrected}
    filled: list[dict] = []
    for previous, current in zip(corrected, corrected[1:]):
        filled.append(previous)
        prev_template = previous.get("template") or {}
        current_template = current.get("template") or {}
        if prev_template.get("sheet_name") != current_template.get("sheet_name"):
            continue
        prev_row = prev_template.get("source_row")
        current_row = current_template.get("source_row")
        if not isinstance(prev_row, int) or not isinstance(current_row, int):
            continue
        if current_row <= prev_row or current_row - prev_row > max_gap_rows + 1:
            continue

        for template in by_sheet.get(prev_template.get("sheet_name"), []):
            source_row = template.get("source_row")
            key = _template_key(template)
            if prev_row < source_row < current_row and key not in existing_keys:
                logger.info(
                    "Filling missing template row %s seq %s between source rows %s and %s",
                    template.get("operation_number"),
                    template.get("sequence"),
                    prev_row,
                    current_row,
                )
                filled.append(_template_enriched_row(template, previous, model))
                existing_keys.add(key)
    filled.append(corrected[-1])
    return filled


def _fill_template_span(corrected: list[dict], template_rows: list[dict], model: str) -> list[dict]:
    """Fill every missing printed template row inside the detected sheet span."""
    if not corrected:
        return corrected

    model_templates = [
        row for row in template_rows
        if row.get("model") == model and row.get("source_row") is not None
    ]
    by_sheet: dict[str, list[dict]] = {}
    for template in model_templates:
        by_sheet.setdefault(template.get("sheet_name"), []).append(template)
    for sheet_rows in by_sheet.values():
        sheet_rows.sort(key=lambda row: (row.get("source_row") or 0, row.get("sequence") or 0))

    rows_by_sheet: dict[str, list[dict]] = {}
    for row in corrected:
        template = row.get("template") or {}
        sheet_name = template.get("sheet_name")
        source_row = template.get("source_row")
        if sheet_name and isinstance(source_row, int):
            rows_by_sheet.setdefault(sheet_name, []).append(row)

    additions: list[dict] = []
    existing_keys = {_template_key(row.get("template")) for row in corrected}
    for sheet_name, sheet_rows in rows_by_sheet.items():
        detected_source_rows = [
            row.get("template", {}).get("source_row")
            for row in sheet_rows
            if isinstance(row.get("template", {}).get("source_row"), int)
        ]
        if not detected_source_rows:
            continue
        start_row = min(detected_source_rows)
        end_row = max(detected_source_rows)
        base = sheet_rows[0]
        for template in by_sheet.get(sheet_name, []):
            source_row = template.get("source_row")
            key = _template_key(template)
            if start_row <= source_row <= end_row and key not in existing_keys:
                logger.info(
                    "Filling missing printed template row %s seq %s inside source row span %s-%s",
                    template.get("operation_number"),
                    template.get("sequence"),
                    start_row,
                    end_row,
                )
                additions.append(_template_enriched_row(template, base, model))
                existing_keys.add(key)

    if not additions:
        return corrected
    combined = [*corrected, *additions]
    return sorted(
        combined,
        key=lambda row: (
            (row.get("template") or {}).get("sheet_name") or "",
            (row.get("template") or {}).get("source_row") or 0,
            (row.get("template") or {}).get("sequence") or 0,
            row.get("id") or 0,
        ),
    )


def apply_standard_template(
    ocr_rows: list[dict],
    db: Session,
    preferred_model: str | None = None,
) -> tuple[list[dict], str | None]:
    """Return rows whose printed fields come only from the template DB.

    Handwritten fields still come from OCR. If a row cannot be matched to a
    template operation, it is not saved as a guessed printed row.
    """
    if not ocr_rows:
        return ocr_rows, None

    template_rows = _load_db_template_rows(db)
    model = _choose_model(ocr_rows, template_rows, preferred_model)
    if not model:
        logger.warning("No standard template model available; refusing to save guessed printed rows.")
        return [], None

    by_op: dict[str, list[dict]] = {}
    op_alias_to_template_op: dict[str, str] = {}
    for template in template_rows:
        if template["model"] != model:
            continue
        template_op = template["operation_number"]
        by_op.setdefault(template_op, []).append(template)
        for alias in _op_aliases(template_op):
            op_alias_to_template_op.setdefault(alias, template_op)

    def resolve_template_op(value) -> str | None:
        op = _clean_op(value)
        if not op:
            return None
        return op if op in by_op else op_alias_to_template_op.get(op)

    corrected: list[dict] = []
    index = 0
    while index < len(ocr_rows):
        row = ocr_rows[index]
        template_op = resolve_template_op(row.get("operation_number"))
        candidates = by_op.get(template_op or "")
        if not template_op or not candidates:
            logger.warning(
                "Dropping OCR row with unmatched printed operation %r; printed fields must come from template.",
                row.get("operation_number"),
            )
            index += 1
            continue

        group = [row]
        index += 1
        while index < len(ocr_rows) and resolve_template_op(ocr_rows[index].get("operation_number")) == template_op:
            group.append(ocr_rows[index])
            index += 1

        measurements: list[float] = []
        measurement_scores: list[float] = []
        for item in group:
            measurements.extend(item.get("measurements") or [])
            scores = (item.get("confidence_scores") or {}).get("measurements") or []
            measurement_scores.extend(scores[: len(item.get("measurements") or [])])

        measurement_offset = 0
        score_offset = 0
        for template_index, template in enumerate(candidates):
            base = group[min(template_index, len(group) - 1)] if group else row
            quantity = template.get("quantity")
            take = max(0, int(quantity or 0))
            template_measurements = measurements[measurement_offset: measurement_offset + take]
            template_scores = measurement_scores[score_offset: score_offset + len(template_measurements)]
            measurement_offset += take
            score_offset += len(template_measurements)

            confidence_scores = {**(base.get("confidence_scores") or {})}
            confidence_scores.update({
                "operation_number": 1.0,
                "process_name": 1.0,
                "quantity": 1.0,
            })
            if template_measurements:
                confidence_scores["measurements"] = template_scores or confidence_scores.get("measurements") or []

            enriched = _template_enriched_row(template, base, model, template_measurements)
            enriched["confidence_scores"] = confidence_scores
            corrected.append(enriched)

    corrected = _fill_template_gaps(corrected, template_rows, model)
    corrected = _fill_template_span(corrected, template_rows, model)
    return corrected, model
